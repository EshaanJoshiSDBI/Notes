import openpyxl as xl
wb = xl.load_workbook('temp.xlsx')
ws = wb['India']
print(ws['A2'].value)
print(ws['C2'].value)
print(ws.cell(2,1).value)
print(ws.max_row)
print(ws.max_column)
players = []
# row numbers - 2,3,4,5,6
# indexes - 1,2,3,4,5
row_num = ws.max_row
for i in range(2,row_num + 1):
    players.append(ws.cell(i,1).value)
print(players)